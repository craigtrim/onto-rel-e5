#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Validate and convert the Ground Truth Excel sheet into a clean CSV.
Usage:
    python scripts/validate_ground_truth.py
"""

from openpyxl import load_workbook
import csv
from pathlib import Path
import sys

# ---------------------------------------------------------------------
# Paths (relative to repo root)
# ---------------------------------------------------------------------
ROOT_DIR = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT_DIR / "data" / "raw" / "Ground Truth.xlsx"
OUTPUT_CSV = ROOT_DIR / "data" / "interim" / "Ground Truth.csv"

# ---------------------------------------------------------------------
# Verify input exists
# ---------------------------------------------------------------------
if not INPUT_XLSX.exists():
    sys.exit(f"❌ Input file not found: {INPUT_XLSX}")

# ---------------------------------------------------------------------
# Load workbook / sheet
# ---------------------------------------------------------------------
wb = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
if "Sheet1" not in wb.sheetnames:
    sys.exit("❌ Expected sheet named 'Sheet1' not found in workbook.")
ws = wb["Sheet1"]

# ---------------------------------------------------------------------
# Header validation
# ---------------------------------------------------------------------
header = [str(c.value).strip() if c.value is not None else "" for c in next(
    ws.iter_rows(max_row=1))]
required = ["A", "B", "A->B", "B->A"]
missing = [c for c in required if c not in header]
if missing:
    sys.exit(f"❌ Missing required columns: {missing}")

col_index = {name: header.index(name) for name in required}

# ---------------------------------------------------------------------
# Row validation and conversion
# ---------------------------------------------------------------------
bad_rows = []
rows_out = []

for excel_rownum, row in enumerate(ws.iter_rows(min_row=2), start=2):
    a = row[col_index["A"]].value
    b = row[col_index["B"]].value
    a2b = row[col_index["A->B"]].value
    b2a = row[col_index["B->A"]].value

    a = str(a).strip() if a is not None else ""
    b = str(b).strip() if b is not None else ""

    def normalize_bool(v, col):
        if isinstance(v, bool):
            return v
        if isinstance(v, str):
            s = v.strip().upper()
            if s in ("TRUE", "T", "YES", "1"):
                return True
            if s in ("FALSE", "F", "NO", "0"):
                return False
        bad_rows.append((excel_rownum, f"invalid {col}", a, b, v))
        return None

    if not a or not b:
        bad_rows.append((excel_rownum, "blank A/B", a, b, (a2b, b2a)))
        continue

    a2b_val = normalize_bool(a2b, "A->B")
    b2a_val = normalize_bool(b2a, "B->A")

    if a2b_val is not None and b2a_val is not None:
        rows_out.append([a, b, str(a2b_val), str(b2a_val)])

# ---------------------------------------------------------------------
# Report and write output
# ---------------------------------------------------------------------
if bad_rows:
    print(f"\n❌ Found {len(bad_rows)} invalid rows. Showing first 20:\n")
    print(f"{'Row':<6} | {'Reason':<14} | {'A':<35} | {'B':<35} | {'Value':<15}")
    print("-" * 115)
    for r in bad_rows[:20]:
        rownum, reason, a, b, val = r
        print(f"{rownum:<6} | {reason:<14} | {a:<35} | {b:<35} | {str(val):<15}")
    sys.exit("\n⚠️  Fix these cells in Excel and rerun.")
else:
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(required)
        writer.writerows(rows_out)

    print(f"✅ All rows passed validation.")
    print(f"💾 CSV written to: {OUTPUT_CSV}")
    print(f"📊 Rows written: {len(rows_out):,}")
